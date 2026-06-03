"""Excel workbook generation for parsed board meeting outcomes."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Announcement, FinancialData
from pdf_parser import FINANCIAL_ROWS
from utils import normalize_date, sanitize_filename

META_COLUMNS = [
    "Source",
    "Company Name",
    "Symbol / Scrip Code",
    "Announcement Date & Time",
    "Subject",
    "PDF URL",
    "PDF File",
    "Screenshots",
    "Parser Status",
    "Parser Message",
    "Meeting Date",
    "Board Meeting Start Time",
    "Board Meeting End Time",
    "Dividend Declared",
    "Dividend Per Share",
    "Dividend",
    "Currency Unit",
    "Validation Status",
    "Validation Errors",
    "Extraction Layer",
    "Parser Layers",
    "Language",
    "Document Type",
    "LLM Status",
]

DEFAULT_PERIOD_COLUMNS = [
    "Q4 FY26",
    "Q3 FY26",
    "Change (in %)",
    "Q4 FY25",
    "Change (in %)",
    "FY26",
    "FY25",
    "Change (in %)",
]

TARGET_METRICS = [
    "Revenue",
    "Expenses",
    "Cost of materials consumed",
    "Change in inventory",
    "Gross Profit",
    "Gross Profit Margin",
    "Employee Benefit Expense",
    "Other expenses",
    "Total Expenses",
    "EBITDA",
    "EBITDA Margin",
    "Depreciation",
    "Finance Cost",
    "Profit before Exceptional Items, Other Income",
    "Other Income",
    "Exceptional items (Discontinued Operations)",
    "Profit Before Tax",
    "Tax Expenses",
    "PAT",
    "PAT Margin",
    "EPS (Basic)",
]


def write_excel(
    records: list[tuple[Announcement, FinancialData]],
    run_date: date,
    output_path: Path | None = None,
) -> Path:
    """Write announcement metadata and parsed financials to an Excel workbook."""

    output_dir = Path("output")
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_path or output_dir / f"board_meeting_outcomes_{run_date.isoformat()}.xlsx"

    workbook = Workbook()
    result_summary = workbook.active
    result_summary.title = "Result Summary"
    _write_result_summary_sheet(result_summary, records)
    outcomes = workbook.create_sheet("Outcomes")
    _write_summary_sheet(outcomes, records)
    _write_financial_table_sheet(workbook, records)
    qa = workbook.create_sheet("QA")
    _write_quality_sheet(qa, records)
    confidence = workbook.create_sheet("Confidence")
    _write_confidence_sheet(confidence, records)
    confidence.sheet_state = "hidden"
    coverage = workbook.create_sheet("Coverage")
    _write_coverage_sheet(coverage, records)
    workbook.save(target)
    return target


def write_alert_excel(
    announcement: Announcement,
    financials: FinancialData,
    extracted_at: datetime,
    output_dir: Path = Path("output") / "excel",
) -> Path:
    """Write one Telegram-ready Excel workbook for a single announcement."""

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = extracted_at.strftime("%d-%m-%Y_%H-%M")
    filename = f"{announcement.source.upper()}_{sanitize_filename(announcement.company_name)}_{timestamp}.xlsx"
    target = output_dir / filename
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Result Summary"

    confidence_score = _estimated_accuracy(financials)
    normalized = _build_normalized_summary(financials)
    source = announcement.source.upper()
    meeting_date = financials.meeting_date or normalize_date(announcement.announcement_datetime)
    alert_columns = _alert_excel_columns(normalized)

    sheet.append([announcement.company_name, source, announcement.identifier])
    sheet.append(["Board Meeting Date", meeting_date, "Extracted At", extracted_at.strftime("%d-%m-%Y %H:%M:%S"), "Parser Status", financials.parser_status])
    headers = ["Rs in Cr"] + [header for header, _ in alert_columns] + ["Source", "Confidence Score", "Extracted At"]
    sheet.append(headers)

    for metric in TARGET_METRICS:
        values = normalized.get(metric, {})
        row_values = [metric]
        for _, getter in alert_columns:
            row_values.append(getter(values, metric))
        row_values.extend([source, confidence_score, extracted_at.strftime("%d-%m-%Y %H:%M:%S")])
        sheet.append(row_values)

    sheet.append([])
    sheet.append(["Dividend Declared", financials.dividend_declared])
    sheet.append(["Dividend Per Share", financials.dividend_per_share])
    sheet.append(["Dividend Text", financials.dividend])
    sheet.append(["PDF URL", announcement.pdf_url])
    sheet.append(["PDF File", str(announcement.pdf_path or "")])
    sheet.append(["Parser Message", financials.parser_message])

    confidence = workbook.create_sheet("Confidence")
    confidence.append(["Metric", "Period", "Confidence"])
    for metric in TARGET_METRICS:
        for period in [item for item in DEFAULT_PERIOD_COLUMNS if not item.startswith("Change")]:
            confidence.append([metric, period, _build_normalized_confidence_summary(financials).get(metric, {}).get(period, 0.0)])
    confidence.sheet_state = "hidden"

    dark_fill = PatternFill("solid", fgColor="1A1A2E")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    key_fill = PatternFill("solid", fgColor="1F4E16")
    for cell in sheet[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=4, max_row=3 + len(TARGET_METRICS)):
        metric = str(row[0].value or "")
        row[0].font = Font(bold=True, color="FFFFFF" if metric in _highlight_metrics() else "000000")
        row[0].fill = key_fill if metric in _highlight_metrics() else green_fill
        for cell in row[1 : 1 + len(alert_columns)]:
            cell.alignment = Alignment(horizontal="right")
    confidence_column = 2 + len(alert_columns) + 1
    extracted_at_column = confidence_column + 1
    sheet.column_dimensions[get_column_letter(confidence_column)].hidden = True
    sheet.column_dimensions[get_column_letter(extracted_at_column)].hidden = True
    sheet.freeze_panes = "B4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(sheet.max_column)}{3 + len(TARGET_METRICS)}"
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
        sheet.column_dimensions[letter].width = max(width, 12)
    workbook.save(target)
    return target


def _alert_excel_columns(normalized: dict[str, dict[str, str]]) -> list[tuple[str, object]]:
    """Return only alert workbook columns with values for this PDF."""

    def period_has_value(period: str) -> bool:
        return any(values.get(period) not in ("", None) for values in normalized.values())

    def change_has_value(current: str, previous: str) -> bool:
        for metric, values in normalized.items():
            if "Margin" in metric:
                continue
            if _change_percent(values.get(current, ""), values.get(previous, "")):
                return True
        return False

    columns: list[tuple[str, object]] = []
    period_presence = {
        "Q4 FY26": period_has_value("Q4 FY26"),
        "Q3 FY26": period_has_value("Q3 FY26"),
        "Q4 FY25": period_has_value("Q4 FY25"),
        "FY26": period_has_value("FY26"),
        "FY25": period_has_value("FY25"),
    }
    if period_presence["Q4 FY26"]:
        columns.append(("Q4 FY26", lambda values, _metric, period="Q4 FY26": values.get(period, "")))
    if period_presence["Q3 FY26"]:
        columns.append(("Q3 FY26", lambda values, _metric, period="Q3 FY26": values.get(period, "")))
    if period_presence["Q4 FY26"] and period_presence["Q3 FY26"] and change_has_value("Q4 FY26", "Q3 FY26"):
        columns.append(
            (
                "Change (in %)",
                lambda values, metric, current="Q4 FY26", previous="Q3 FY26": ""
                if "Margin" in metric
                else _change_percent(values.get(current, ""), values.get(previous, "")),
            )
        )
    if period_presence["Q4 FY25"]:
        columns.append(("Q4 FY25", lambda values, _metric, period="Q4 FY25": values.get(period, "")))
    if period_presence["Q4 FY26"] and period_presence["Q4 FY25"] and change_has_value("Q4 FY26", "Q4 FY25"):
        columns.append(
            (
                "Change (in %)",
                lambda values, metric, current="Q4 FY26", previous="Q4 FY25": ""
                if "Margin" in metric
                else _change_percent(values.get(current, ""), values.get(previous, "")),
            )
        )
    if period_presence["FY26"]:
        columns.append(("FY26", lambda values, _metric, period="FY26": values.get(period, "")))
    if period_presence["FY25"]:
        columns.append(("FY25", lambda values, _metric, period="FY25": values.get(period, "")))
    if period_presence["FY26"] and period_presence["FY25"] and change_has_value("FY26", "FY25"):
        columns.append(
            (
                "Change (in %)",
                lambda values, metric, current="FY26", previous="FY25": ""
                if "Margin" in metric
                else _change_percent(values.get(current, ""), values.get(previous, "")),
            )
        )
    return columns


def _write_result_summary_sheet(sheet, records: list[tuple[Announcement, FinancialData]]) -> None:
    """Write normalized financial summaries matching the requested screenshot layout."""

    current_row = 1
    for announcement, financials in records:
        normalized = _build_normalized_summary(financials)
        sheet.cell(current_row, 1, announcement.company_name).font = Font(bold=True, size=14)
        sheet.cell(current_row, 2, announcement.source)
        sheet.cell(current_row, 3, announcement.identifier)
        current_row += 1

        headers = ["Rs in Cr"] + DEFAULT_PERIOD_COLUMNS
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(current_row, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for metric in TARGET_METRICS:
            values = normalized.get(metric, {})
            sheet.cell(current_row, 1, metric)
            sheet.cell(current_row, 1).font = Font(bold=True, color="FFFFFF" if metric in _highlight_metrics() else "000000")
            sheet.cell(current_row, 1).fill = PatternFill(
                "solid",
                fgColor="1F4E16" if metric in _highlight_metrics() else "C6E0B4",
            )
            q4 = values.get("Q4 FY26", "")
            q3 = values.get("Q3 FY26", "")
            q4_prev = values.get("Q4 FY25", "")
            fy = values.get("FY26", "")
            fy_prev = values.get("FY25", "")
            row_values = [
                q4,
                q3,
                "" if "Margin" in metric else _change_percent(q4, q3),
                q4_prev,
                "" if "Margin" in metric else _change_percent(q4, q4_prev),
                fy,
                fy_prev,
                "" if "Margin" in metric else _change_percent(fy, fy_prev),
            ]
            for col_idx, value in enumerate(row_values, start=2):
                sheet.cell(current_row, col_idx, value)
            current_row += 1
        current_row += 2
    _style_sheet(sheet)


def _highlight_metrics() -> set[str]:
    """Return rows styled as key result rows."""

    return {
        "Revenue",
        "Gross Profit",
        "Gross Profit Margin",
        "EBITDA",
        "EBITDA Margin",
        "Profit before Exceptional Items, Other Income",
        "Profit Before Tax",
        "PAT",
        "PAT Margin",
    }


def _build_normalized_summary(financials: FinancialData) -> dict[str, dict[str, str]]:
    """Map parsed period labels into Q4/Q3/FY result-summary columns."""

    mapped: dict[str, dict[str, float]] = {}
    for metric, period_values in financials.rows.items():
        canonical = _canonical_metric(metric)
        for period, raw_value in period_values.items():
            target_period = _target_period(period)
            value = _to_float(raw_value)
            if not canonical or not target_period or value is None:
                continue
            mapped.setdefault(canonical, {})[target_period] = value

    _derive_summary_values(mapped)
    return {
        metric: {period: _format_value(value, metric) for period, value in values.items()}
        for metric, values in mapped.items()
    }


def _canonical_metric(metric: str) -> str:
    """Map parser metric names to the requested output metric names."""

    aliases = {
        "Revenue from operations": "Revenue",
        "Total Income": "Revenue",
        "Profit before tax and exceptional items": "Profit before Exceptional Items, Other Income",
        "Profit for the period/year": "PAT",
        "Current Tax": "Tax Expenses",
        "Deferred Tax": "Tax Expenses",
        "EPS (Diluted)": "EPS (Basic)",
    }
    return aliases.get(metric, metric)


def _target_period(period: str) -> str:
    """Map PDF period labels into Q4 FY26/Q3 FY26/Q4 FY25/FY26/FY25."""

    cleaned = (period or "").lower()
    if re.search(r"q4\s*fy\s*26", cleaned):
        return "Q4 FY26"
    if re.search(r"q3\s*fy\s*26", cleaned):
        return "Q3 FY26"
    if re.search(r"q4\s*fy\s*25", cleaned):
        return "Q4 FY25"
    if re.fullmatch(r"fy\s*26", cleaned.replace(" ", "")):
        return "FY26"
    if re.fullmatch(r"fy\s*25", cleaned.replace(" ", "")):
        return "FY25"

    year_match = re.search(r"20(\d{2})", cleaned)
    if not year_match:
        return ""
    fy_suffix = year_match.group(1)
    if "year ended" in cleaned:
        return f"FY{fy_suffix}"
    if "31 march" in cleaned or "31-mar" in cleaned or "31/03" in cleaned:
        return f"Q4 FY{fy_suffix}"
    if "31 december" in cleaned or "31-dec" in cleaned or "31/12" in cleaned:
        return f"Q3 FY{int(fy_suffix) + 1:02d}"
    if "30 september" in cleaned or "30-sep" in cleaned or "30/09" in cleaned:
        return f"Q2 FY{int(fy_suffix) + 1:02d}"
    if "30 june" in cleaned or "30-jun" in cleaned or "30/06" in cleaned:
        return f"Q1 FY{int(fy_suffix) + 1:02d}"
    return ""


def _derive_summary_values(mapped: dict[str, dict[str, float]]) -> None:
    """Derive requested rows such as Gross Profit, EBITDA, and margins when possible."""

    all_periods = set().union(*(values.keys() for values in mapped.values())) if mapped else set()
    for period in all_periods:
        revenue = _get(mapped, "Revenue", period)
        cost = _get(mapped, "Cost of materials consumed", period)
        inventory = _get(mapped, "Change in inventory", period)
        if _missing(mapped, "Gross Profit", period) and revenue is not None and cost is not None and inventory is not None:
            mapped.setdefault("Gross Profit", {})[period] = revenue - cost - inventory
        gross_profit = _get(mapped, "Gross Profit", period)
        if _missing(mapped, "Gross Profit Margin", period) and revenue not in (None, 0) and gross_profit is not None:
            mapped.setdefault("Gross Profit Margin", {})[period] = gross_profit / revenue * 100

        if _missing(mapped, "Tax Expenses", period):
            current_tax = _get(mapped, "Current Tax", period)
            deferred_tax = _get(mapped, "Deferred Tax", period)
            if current_tax is not None or deferred_tax is not None:
                mapped.setdefault("Tax Expenses", {})[period] = (current_tax or 0) + (deferred_tax or 0)

        pbt = _get(mapped, "Profit Before Tax", period)
        finance = _get(mapped, "Finance Cost", period) or 0
        depreciation = _get(mapped, "Depreciation", period) or 0
        other_income = _get(mapped, "Other Income", period) or 0
        exceptional = _get(mapped, "Exceptional items (Discontinued Operations)", period) or 0
        if _missing(mapped, "EBITDA", period) and pbt is not None:
            mapped.setdefault("EBITDA", {})[period] = pbt + finance + depreciation - other_income - exceptional
        ebitda = _get(mapped, "EBITDA", period)
        if _missing(mapped, "EBITDA Margin", period) and revenue not in (None, 0) and ebitda is not None:
            mapped.setdefault("EBITDA Margin", {})[period] = ebitda / revenue * 100
        pat = _get(mapped, "PAT", period)
        if _missing(mapped, "PAT Margin", period) and revenue not in (None, 0) and pat is not None:
            mapped.setdefault("PAT Margin", {})[period] = pat / revenue * 100


def _get(mapped: dict[str, dict[str, float]], metric: str, period: str) -> float | None:
    """Get one numeric mapped value."""

    return mapped.get(metric, {}).get(period)


def _missing(mapped: dict[str, dict[str, float]], metric: str, period: str) -> bool:
    """Return whether a metric-period value is missing."""

    return period not in mapped.get(metric, {})


def _to_float(value: object) -> float | None:
    """Parse numeric cell text into float."""

    if value is None:
        return None
    text = str(value).strip().replace(",", "").replace(" ", "")
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    is_percent = text.endswith("%")
    text = text.rstrip("%")
    try:
        parsed = float(text)
    except ValueError:
        return None
    return -parsed if negative else parsed


def _format_value(value: float, metric: str) -> str:
    """Format a summary numeric value."""

    if "Margin" in metric:
        return f"{value:.2f}%"
    if metric.startswith("EPS"):
        return f"{value:.2f}"
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _change_percent(current: str, previous: str) -> str:
    """Calculate percentage change using absolute prior-period denominator."""

    current_value = _to_float(current)
    previous_value = _to_float(previous)
    if current_value is None or previous_value in (None, 0):
        return ""
    return f"{((current_value - previous_value) / abs(previous_value) * 100):.2f}%"


def _write_summary_sheet(sheet, records: list[tuple[Announcement, FinancialData]]) -> None:
    """Write one row per announcement with flattened financial columns."""

    periods = _collect_periods(records)
    headers = META_COLUMNS + [f"{row} - {period}" for row in FINANCIAL_ROWS for period in periods]
    sheet.append(headers)
    for announcement, financials in records:
        _validate_record(announcement, financials)
        row = [
            announcement.source,
            announcement.company_name,
            announcement.identifier,
            announcement.announcement_datetime,
            announcement.subject,
            announcement.pdf_url,
            str(announcement.pdf_path or ""),
            "; ".join(financials.screenshots),
            financials.parser_status,
            financials.parser_message,
            financials.meeting_date,
            financials.board_meeting_start_time,
            financials.board_meeting_end_time,
            financials.dividend_declared,
            financials.dividend_per_share,
            financials.dividend,
            financials.currency_unit,
            financials.validation_status,
            "; ".join(financials.validation_errors),
            financials.extraction_layer,
            "; ".join(financials.parser_layers),
            financials.language,
            financials.document_type,
            financials.llm_status,
        ]
        for metric in FINANCIAL_ROWS:
            for period in periods:
                row.append(financials.rows.get(metric, {}).get(period, ""))
        sheet.append(row)
    _style_sheet(sheet)


def _write_financial_table_sheet(
    workbook: Workbook,
    records: list[tuple[Announcement, FinancialData]],
) -> None:
    """Write a screenshot-like financial statement layout per announcement."""

    sheet = workbook.create_sheet("Financial Tables")
    current_row = 1
    for announcement, financials in records:
        periods = financials.periods or DEFAULT_PERIOD_COLUMNS
        sheet.cell(current_row, 1, announcement.company_name).font = Font(bold=True, size=14)
        sheet.cell(current_row, 2, announcement.source)
        sheet.cell(current_row, 3, announcement.identifier)
        current_row += 1

        header = [financials.currency_unit or "Rs in Cr"] + periods
        for col_idx, value in enumerate(header, start=1):
            cell = sheet.cell(current_row, col_idx, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for metric in FINANCIAL_ROWS:
            sheet.cell(current_row, 1, metric)
            sheet.cell(current_row, 1).font = Font(bold=True)
            sheet.cell(current_row, 1).fill = PatternFill("solid", fgColor="C6E0B4")
            for col_idx, period in enumerate(periods, start=2):
                sheet.cell(current_row, col_idx, financials.rows.get(metric, {}).get(period, ""))
            current_row += 1
        current_row += 2
    _style_sheet(sheet)


def _collect_periods(records: list[tuple[Announcement, FinancialData]]) -> list[str]:
    """Collect all period labels found across parsed PDFs."""

    periods: list[str] = []
    for _, financials in records:
        for period in financials.periods:
            if period and period not in periods:
                periods.append(period)
    return periods or DEFAULT_PERIOD_COLUMNS


def _style_sheet(sheet) -> None:
    """Apply basic widths, frozen panes, filters, and header styling."""

    sheet.freeze_panes = "A2"
    max_column = sheet.max_column
    max_row = sheet.max_row
    if max_row and max_column:
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
        sheet.column_dimensions[letter].width = max(width, 12)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_quality_sheet(sheet, records: list[tuple[Announcement, FinancialData]]) -> None:
    """Write extraction quality diagnostics and confidence estimates."""

    headers = [
        "Source",
        "Company Name",
        "Identifier",
        "PDF File",
        "Parser Status",
        "Parser Message",
        "Periods Found",
        "Metrics Found",
        "Values Found",
        "Screenshots",
        "Extraction Layer",
        "Parser Layers",
        "Language",
        "Document Type",
        "Validation Status",
        "Validation Errors",
        "LLM Status",
        "Estimated Accuracy %",
    ]
    sheet.append(headers)
    for announcement, financials in records:
        values_found = sum(len(values) for values in financials.rows.values())
        metrics_found = len(financials.rows)
        periods_found = len(financials.periods)
        sheet.append(
            [
                announcement.source,
                announcement.company_name,
                announcement.identifier,
                str(announcement.pdf_path or ""),
                financials.parser_status,
                financials.parser_message,
                periods_found,
                metrics_found,
                values_found,
                "; ".join(financials.screenshots),
                financials.extraction_layer,
                "; ".join(financials.parser_layers),
                financials.language,
                financials.document_type,
                financials.validation_status,
                "; ".join(financials.validation_errors),
                financials.llm_status,
                _estimated_accuracy(financials),
            ]
        )
    _style_sheet(sheet)


def _write_confidence_sheet(sheet, records: list[tuple[Announcement, FinancialData]]) -> None:
    """Write field-level confidence scores for audit."""

    periods = [period for period in DEFAULT_PERIOD_COLUMNS if not period.startswith("Change")]
    headers = ["Source", "Company Name", "PDF File"] + [f"{row} - {period}" for row in TARGET_METRICS for period in periods]
    sheet.append(headers)
    for announcement, financials in records:
        row = [announcement.source, announcement.company_name, str(announcement.pdf_path or "")]
        confidence_map = _build_normalized_confidence_summary(financials)
        for metric in TARGET_METRICS:
            for period in periods:
                row.append(confidence_map.get(metric, {}).get(period, 0.0))
        sheet.append(row)
    if sheet.max_column > 3:
        sheet.column_dimensions.group("D", get_column_letter(sheet.max_column), hidden=True)
    _style_sheet(sheet)


def _write_coverage_sheet(sheet, records: list[tuple[Announcement, FinancialData]]) -> None:
    """Write per-field coverage and overall non-empty cell coverage."""

    periods = [period for period in DEFAULT_PERIOD_COLUMNS if not period.startswith("Change")]
    sheet.append(["Metric", "Expected Cells", "Non-empty Cells", "Coverage %", "Average Confidence"])
    total_expected = 0
    total_non_empty = 0
    for metric in TARGET_METRICS:
        expected = len(records) * len(periods)
        non_empty = 0
        confidence_total = 0.0
        confidence_count = 0
        for _, financials in records:
            normalized = _build_normalized_summary(financials)
            confidence_map = _build_normalized_confidence_summary(financials)
            for period in periods:
                value = normalized.get(metric, {}).get(period, "")
                if value not in ("", None):
                    non_empty += 1
                    confidence_total += confidence_map.get(metric, {}).get(period, 0.0)
                    confidence_count += 1
        total_expected += expected
        total_non_empty += non_empty
        sheet.append(
            [
                metric,
                expected,
                non_empty,
                round(non_empty / expected * 100, 2) if expected else 0,
                round(confidence_total / confidence_count, 2) if confidence_count else 0,
            ]
        )
    sheet.append([])
    sheet.append(["TOTAL", total_expected, total_non_empty, round(total_non_empty / total_expected * 100, 2) if total_expected else 0, ""])
    _style_sheet(sheet)


def _build_normalized_confidence_summary(financials: FinancialData) -> dict[str, dict[str, float]]:
    """Map raw field confidence to normalized Result Summary periods/metrics."""

    mapped: dict[str, dict[str, float]] = {}
    for metric, period_values in financials.field_confidence.items():
        canonical = _canonical_metric(metric)
        if not canonical:
            continue
        for period, confidence in period_values.items():
            target_period = _target_period(period)
            if not target_period:
                continue
            existing = mapped.setdefault(canonical, {}).get(target_period, 0.0)
            mapped[canonical][target_period] = max(existing, float(confidence or 0.0))
    for metric in ("Gross Profit", "Gross Profit Margin", "EBITDA", "EBITDA Margin", "PAT Margin"):
        for period in DEFAULT_PERIOD_COLUMNS:
            if period.startswith("Change"):
                continue
            if _build_normalized_summary(financials).get(metric, {}).get(period) and not mapped.get(metric, {}).get(period):
                mapped.setdefault(metric, {})[period] = 0.4
    return mapped


def _estimated_accuracy(financials: FinancialData) -> int:
    """Estimate extraction confidence; this is not a ground-truth audit score."""

    values_found = sum(len(values) for values in financials.rows.values())
    metrics_found = len(financials.rows)
    periods_found = len(financials.periods)
    if financials.parser_status in {"scanned_or_empty", "unreadable", "parse_timeout", "parse_error"}:
        return 40
    if financials.parser_status == "no_financial_data":
        if _has_strong_financial_result_evidence(financials):
            return 80
        if financials.parser_message.startswith("Only incidental or suspect"):
            return 100
        return 95
    if financials.parser_status.startswith("parsed") and values_found >= 10 and periods_found >= 3:
        return 100
    if financials.parser_status.startswith("parsed") and values_found >= 5 and periods_found >= 3:
        return 95
    score = 0
    if financials.parser_status.startswith("parsed"):
        score += 35
    if periods_found >= 5:
        score += 20
    elif periods_found >= 3:
        score += 12
    elif periods_found >= 1:
        score += 5
    score += min(metrics_found * 2, 20)
    score += min(values_found, 25)
    return max(0, min(score, 100))


def _has_strong_financial_result_evidence(financials: FinancialData) -> bool:
    """Return whether a no-data PDF appears to contain a result table the parser missed."""

    text = (financials.text_excerpt or "").lower()
    strong_markers = (
        "standalone financial results",
        "consolidated financial results",
        "audited financial results",
        "unaudited financial results",
        "statement of financial results",
        "quarter and year ended",
        "quarter ended and year ended",
    )
    return any(marker in text for marker in strong_markers)


def _validate_record(announcement: Announcement, financials: FinancialData) -> None:
    """Validate one announcement/financial extraction before workbook write."""

    errors: list[str] = []
    if not announcement.company_name.strip():
        errors.append("Company Name is empty")
    date_value = financials.meeting_date or announcement.announcement_datetime
    normalized_date = normalize_date(str(date_value))
    if not normalized_date or not re.search(r"\d{2}-\d{2}-\d{4}", normalized_date):
        errors.append("Date is not parseable as DD-MM-YYYY")
    for metric, period_values in financials.rows.items():
        for period, value in period_values.items():
            parsed = _to_float(value)
            if parsed is None:
                errors.append(f"{metric} {period} is not numeric")
                continue
            if metric.startswith("EPS") and not (-1000 <= parsed <= 10000):
                errors.append(f"{metric} {period} EPS outside sanity range")
    if financials.dividend_per_share:
        dividend_value = _to_float(financials.dividend_per_share)
        if dividend_value is None or dividend_value < 0:
            errors.append("Dividend Per Share is invalid")
    if financials.dividend_declared.lower() == "yes" and not financials.dividend_per_share:
        errors.append("Dividend Declared is Yes but Dividend Per Share is blank")
    financials.validation_errors = errors
    financials.validation_status = "VALIDATION_ERROR" if errors else "OK"
